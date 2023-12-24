import os

from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
def excel_to_pdf(excel_path,pdf_path):


    workbook=load_workbook(excel_path)
    print("book",workbook)
    worksheet= workbook.active
    max_row=worksheet.max_row
    max_column=worksheet.max_column
    print(max_column,"rows",max_row)
    c=canvas.Canvas(pdf_path,pagesize=landscape(letter))
    top_margin=1*inch
    left_margin=0.25*inch
    bottom_margin=0.25*inch
    right_margin=0.25*inch
    cell_width=(11*inch - left_margin -right_margin)/max_column
    cell_height=(11*inch - top_margin -bottom_margin)/max_row

    for row in range(1,max_row+1):
        for column in range(1,max_column+1):
            cell=worksheet.cell(row=row,column=column)
            text=str(cell.value)
            print(cell.value)
            x=left_margin+(column-1)*cell_width
            y=11*inch - (top_margin+row * cell_height)
            print(int(x),int(y))
            c.drawString(x, y, text)

    #
    c.save()


if __name__=="__main__":
    excel_path=os.path.abspath("cd.xlsx")
    pdf_path=os.path.abspath("output.pdf")
    excel_to_pdf(excel_path,pdf_path)